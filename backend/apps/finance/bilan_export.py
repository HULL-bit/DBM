"""
Bilan financier par catégorie de charge (Magal, Gamou, Kazu Rajabb, Koor, Social, Xelcom,
Mensualités, Autres) : montant collecté (cotisations payées), dépenses validées, reste.
Export en Excel ou PDF.
"""
from decimal import Decimal

from django.db.models import Sum

from .models import CotisationMensuelle, Depense, CATEGORIES_CHARGE


def calculer_bilan(annee=None):
    """Retourne une ligne par catégorie + une ligne total.

    Les catégories de dépense sont un choix fermé (CATEGORIES_CHARGE), mais l'objet d'une
    assignation (cotisation) est un texte libre : un objet personnalisé (ex. "Tabaski",
    "Social Segn Fallou") ne correspond à aucun des codes fixes. Sans le détecter ici, son
    montant collecté disparaissait purement et simplement du bilan (et donc du total). On
    ajoute donc une ligne pour chaque objet réellement utilisé, en plus des catégories connues.
    """
    lignes = []
    total_collecte = Decimal('0')
    total_depense = Decimal('0')

    cotisations_payees = CotisationMensuelle.objects.filter(statut='payee')
    depenses_validees = Depense.objects.filter(statut='validee')
    if annee:
        cotisations_payees = cotisations_payees.filter(annee=annee)
        depenses_validees = depenses_validees.filter(date_depense__year=annee)

    codes_connus = [code for code, _ in CATEGORIES_CHARGE]
    libelles_connus = dict(CATEGORIES_CHARGE)
    codes_personnalises = sorted({
        v.strip().upper()
        for v in cotisations_payees.filter(type_cotisation='assignation')
        .exclude(objet_assignation__iexact='')
        .values_list('objet_assignation', flat=True)
        .distinct()
        if v.strip().upper() not in codes_connus
    })

    for code in [*codes_connus, *codes_personnalises]:
        libelle = libelles_connus.get(code, code.title())
        if code == 'MENSUALITE':
            cotisations = cotisations_payees.filter(type_cotisation='mensualite')
        else:
            cotisations = cotisations_payees.filter(type_cotisation='assignation', objet_assignation__iexact=code)
        collecte = cotisations.aggregate(total=Sum('montant'))['total'] or Decimal('0')

        # Sans objet pour une dépense (categorie est un choix fermé) : ne matche que les
        # codes connus, ce qui est correct puisqu'une dépense ne peut pas avoir un objet
        # personnalisé.
        depense = depenses_validees.filter(categorie=code).aggregate(total=Sum('montant'))['total'] or Decimal('0')

        lignes.append({
            'categorie': code,
            'categorie_display': libelle,
            'montant_collecte': collecte,
            'montant_depense': depense,
            'reste': collecte - depense,
        })
        total_collecte += collecte
        total_depense += depense

    lignes.append({
        'categorie': 'TOTAL',
        'categorie_display': 'Total',
        'montant_collecte': total_collecte,
        'montant_depense': total_depense,
        'reste': total_collecte - total_depense,
    })
    return lignes


def export_bilan_excel(annee=None):
    try:
        import openpyxl
        from openpyxl.styles import Font, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return None

    lignes = calculer_bilan(annee)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bilan financier"
    header_font = Font(bold=True)
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.cell(row=1, column=1, value="Bilan financier — Daara Barakatul Mahaahidi").font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=f"Année : {annee}" if annee else "Toutes années")

    headers = ['Catégorie', 'Montant collecté (FCFA)', 'Dépenses (FCFA)', 'Reste (FCFA)']
    row0 = 4
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row0, column=col, value=h)
        c.font = header_font
        c.border = border

    for i, l in enumerate(lignes, row0 + 1):
        est_total = l['categorie'] == 'TOTAL'
        c1 = ws.cell(row=i, column=1, value=l['categorie_display'])
        c2 = ws.cell(row=i, column=2, value=float(l['montant_collecte']))
        c3 = ws.cell(row=i, column=3, value=float(l['montant_depense']))
        c4 = ws.cell(row=i, column=4, value=float(l['reste']))
        for c in (c1, c2, c3, c4):
            c.border = border
            if est_total:
                c.font = header_font
    for col in range(1, 5):
        ws.column_dimensions[get_column_letter(col)].width = 24

    from io import BytesIO
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_bilan_pdf(annee=None):
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
    except ImportError:
        return None

    lignes = calculer_bilan(annee)

    from io import BytesIO
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, rightMargin=2 * cm, leftMargin=2 * cm, topMargin=2 * cm, bottomMargin=2 * cm)
    elements = []

    periode_str = f"Année {annee}" if annee else "Toutes années"
    from utils.pdf_header import build_pdf_header
    elements.extend(build_pdf_header("Bilan financier", periode_str))

    data = [['Catégorie', 'Collecté (FCFA)', 'Dépenses (FCFA)', 'Reste (FCFA)']]
    for l in lignes:
        data.append([
            l['categorie_display'],
            f"{float(l['montant_collecte']):,.0f}",
            f"{float(l['montant_depense']):,.0f}",
            f"{float(l['reste']):,.0f}",
        ])

    table = Table(data, colWidths=[5 * cm, 4 * cm, 4 * cm, 4 * cm])
    style = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2D5F3F')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#F4EAD5')),
    ]
    table.setStyle(TableStyle(style))
    elements.append(table)

    doc.build(elements)
    buf.seek(0)
    return buf
