"""
Export des rapports de cotisations en PDF ou Excel.
Statistiques globales, par membre (taux, montants), somme totale collectée, etc.
"""
from decimal import Decimal
from collections import defaultdict

from .models import CotisationMensuelle


def _get_queryset(date_debut=None, date_fin=None, annee=None, mois=None):
    """Queryset des cotisations. Filtres optionnels par période."""
    qs = CotisationMensuelle.objects.exclude(statut='annulee').select_related('membre').order_by('annee', 'mois')
    if annee:
        qs = qs.filter(annee=annee)
    if mois:
        qs = qs.filter(mois=mois)
    if date_debut:
        # Filtrer par mois/annee: on considère la date d'échéance ou mois/annee
        from datetime import date
        qs = qs.filter(annee__gte=date_debut.year).exclude(
            annee=date_debut.year, mois__lt=date_debut.month
        )
    if date_fin:
        from datetime import date
        qs = qs.filter(annee__lte=date_fin.year).exclude(
            annee=date_fin.year, mois__gt=date_fin.month
        )
    return qs


def _get_stats_par_membre(qs):
    """Retourne [{membre_id, nom, nb_total, nb_payees, nb_en_attente, nb_retard, montant_total, montant_paye, taux_cotisation}]"""
    from django.db.models import Sum, Q
    membre_data = defaultdict(lambda: {
        'nb_total': 0, 'nb_payees': 0, 'nb_en_attente': 0, 'nb_retard': 0,
        'montant_total': Decimal('0'), 'montant_paye': Decimal('0'), 'nom': ''
    })
    for c in qs:
        mid = c.membre_id
        membre_data[mid]['nom'] = c.membre.get_full_name() if c.membre else f'Membre #{mid}'
        membre_data[mid]['nb_total'] += 1
        membre_data[mid]['montant_total'] += c.montant
        if c.statut == 'payee':
            membre_data[mid]['nb_payees'] += 1
            membre_data[mid]['montant_paye'] += c.montant
        elif c.statut == 'en_attente':
            membre_data[mid]['nb_en_attente'] += 1
        elif c.statut == 'retard':
            membre_data[mid]['nb_retard'] += 1

    result = []
    for mid, d in membre_data.items():
        nb_total = d['nb_total']
        nb_payees = d['nb_payees']
        montant_total = d['montant_total']
        montant_paye = d['montant_paye']
        taux = round(100 * nb_payees / nb_total, 1) if nb_total else 0
        taux_montant = round(float(100 * montant_paye / montant_total), 1) if montant_total else 0
        result.append({
            'membre_id': mid, 'nom': d['nom'],
            'nb_total': nb_total, 'nb_payees': nb_payees,
            'nb_en_attente': d.get('nb_en_attente', 0), 'nb_retard': d.get('nb_retard', 0),
            'montant_total': montant_total, 'montant_paye': montant_paye,
            'montant_restant': montant_total - montant_paye,
            'taux_cotisation': taux, 'taux_montant': taux_montant
        })
    return sorted(result, key=lambda x: (-x['montant_paye'], x['nom']))


def export_rapport_excel(date_debut=None, date_fin=None, annee=None, mois=None):
    """Export Excel des cotisations + feuille statistiques + feuille par membre."""
    try:
        import openpyxl
        from openpyxl.styles import Font, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return None

    qs = _get_queryset(date_debut, date_fin, annee, mois)
    stats_membres = _get_stats_par_membre(qs)

    from django.db.models import Sum, Q
    agg = qs.aggregate(
        total_assigne=Sum('montant'),
        total_paye=Sum('montant', filter=Q(statut='payee')),
    )
    montant_total_collecte = agg.get('total_paye') or Decimal('0')
    montant_total_assigne = agg.get('total_assigne') or Decimal('0')
    nb_total = qs.count()
    nb_payees = qs.filter(statut='payee').count()
    nb_en_attente = qs.filter(statut='en_attente').count()
    nb_retard = qs.filter(statut='retard').count()
    taux_paiement = round(100 * nb_payees / nb_total, 1) if nb_total else 0

    wb = openpyxl.Workbook()
    header_font = Font(bold=True)
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Feuille 1 : Statistiques globales
    ws_stats = wb.active
    ws_stats.title = "Statistiques"
    periode_str = "Toutes les cotisations"
    if annee:
        periode_str = f"Année {annee}" + (f" - Mois {mois}" if mois else "")
    elif date_debut and date_fin:
        periode_str = f"{date_debut.strftime('%d/%m/%Y')} — {date_fin.strftime('%d/%m/%Y')}"

    ws_stats.cell(row=1, column=1, value="Rapport des cotisations").font = Font(bold=True, size=14)
    ws_stats.cell(row=2, column=1, value=f"Période : {periode_str}")
    row_s = 4
    for label, val in [
        ("Nombre total de cotisations", nb_total),
        ("Cotisations payées", nb_payees),
        ("En attente", nb_en_attente),
        ("En retard", nb_retard),
        ("Taux de paiement (%)", f"{taux_paiement}%"),
        ("Montant total assigné (FCFA)", float(montant_total_assigne)),
        ("Somme totale collectée (FCFA)", float(montant_total_collecte)),
        ("Reste à collecter (FCFA)", float(montant_total_assigne - montant_total_collecte)),
    ]:
        ws_stats.cell(row=row_s, column=1, value=label).font = header_font
        ws_stats.cell(row=row_s, column=2, value=val).border = border
        row_s += 1

    # Feuille 2 : Taux et montants par membre
    ws_membres = wb.create_sheet("Taux par membre", 1)
    h_membres = ['Membre', 'Cotisations totales', 'Payées', 'En attente', 'En retard', 'Montant assigné (FCFA)',
                 'Montant payé (FCFA)', 'Reste (FCFA)', 'Taux cotisation (%)', 'Taux montant (%)']
    for col, h in enumerate(h_membres, 1):
        c = ws_membres.cell(row=1, column=col, value=h)
        c.font = header_font
        c.border = border
    for i, s in enumerate(stats_membres, 2):
        ws_membres.cell(row=i, column=1, value=s['nom']).border = border
        ws_membres.cell(row=i, column=2, value=s['nb_total']).border = border
        ws_membres.cell(row=i, column=3, value=s['nb_payees']).border = border
        ws_membres.cell(row=i, column=4, value=s['nb_en_attente']).border = border
        ws_membres.cell(row=i, column=5, value=s['nb_retard']).border = border
        ws_membres.cell(row=i, column=6, value=float(s['montant_total'])).border = border
        ws_membres.cell(row=i, column=7, value=float(s['montant_paye'])).border = border
        ws_membres.cell(row=i, column=8, value=float(s['montant_restant'])).border = border
        ws_membres.cell(row=i, column=9, value=f"{s['taux_cotisation']}%").border = border
        ws_membres.cell(row=i, column=10, value=f"{s['taux_montant']}%").border = border
    for col in range(1, 11):
        ws_membres.column_dimensions[get_column_letter(col)].width = 18

    # Feuille 3 : Détail des cotisations
    ws = wb.create_sheet("Détail cotisations", 2)
    headers = ['Membre', 'Type', 'Objet', 'Mois', 'Année', 'Montant (FCFA)', 'Statut', 'Date échéance', 'Date paiement', 'Mode paiement']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = header_font
        c.border = border
    taux_par_membre = {s['membre_id']: s['taux_cotisation'] for s in stats_membres}
    for i, c in enumerate(qs, 2):
        ws.cell(row=i, column=1, value=c.membre.get_full_name() if c.membre else '').border = border
        ws.cell(row=i, column=2, value=c.get_type_cotisation_display()).border = border
        ws.cell(row=i, column=3, value=c.objet_assignation or '—').border = border
        ws.cell(row=i, column=4, value=c.mois).border = border
        ws.cell(row=i, column=5, value=c.annee).border = border
        ws.cell(row=i, column=6, value=float(c.montant)).border = border
        ws.cell(row=i, column=7, value=c.get_statut_display()).border = border
        ws.cell(row=i, column=8, value=c.date_echeance.strftime('%d/%m/%Y') if c.date_echeance else '').border = border
        ws.cell(row=i, column=9, value=c.date_paiement.strftime('%d/%m/%Y') if c.date_paiement else '').border = border
        mode_disp = {'wave': 'Wave', 'liquide': 'Espèces', 'autre': 'Autre'}.get(c.mode_paiement or 'wave', c.mode_paiement or '')
        ws.cell(row=i, column=10, value=mode_disp).border = border
    for col in range(1, 11):
        ws.column_dimensions[get_column_letter(col)].width = 16

    from io import BytesIO
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_rapport_pdf(date_debut=None, date_fin=None, annee=None, mois=None):
    """Export PDF des cotisations."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    except ImportError:
        return None

    qs = _get_queryset(date_debut, date_fin, annee, mois)
    stats_membres = _get_stats_par_membre(qs)

    from django.db.models import Sum, Q
    agg = qs.aggregate(
        total_assigne=Sum('montant'),
        total_paye=Sum('montant', filter=Q(statut='payee')),
    )
    montant_total_collecte = agg.get('total_paye') or Decimal('0')
    montant_total_assigne = agg.get('total_assigne') or Decimal('0')
    nb_total = qs.count()
    nb_payees = qs.filter(statut='payee').count()
    nb_en_attente = qs.filter(statut='en_attente').count()
    nb_retard = qs.filter(statut='retard').count()
    taux_paiement = round(100 * nb_payees / nb_total, 1) if nb_total else 0

    from io import BytesIO
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    elements = []

    periode_str = "Toutes les cotisations"
    if annee:
        periode_str = f"Année {annee}" + (f" - Mois {mois}" if mois else "")
    elif date_debut and date_fin:
        periode_str = f"{date_debut.strftime('%d/%m/%Y')} — {date_fin.strftime('%d/%m/%Y')}"

    from utils.pdf_header import build_pdf_header
    elements.extend(build_pdf_header("Rapport des cotisations", periode_str))

    # Statistiques
    stats_data = [
        ['Nombre total cotisations', str(nb_total)],
        ['Payées', str(nb_payees)],
        ['En attente', str(nb_en_attente)],
        ['En retard', str(nb_retard)],
        ['Taux paiement (%)', f'{taux_paiement}%'],
        ['Montant total assigné (FCFA)', f'{float(montant_total_assigne):,.0f}'],
        ['Somme totale collectée (FCFA)', f'{float(montant_total_collecte):,.0f}'],
        ['Reste à collecter (FCFA)', f'{float(montant_total_assigne - montant_total_collecte):,.0f}'],
    ]
    stats_table = Table([['Indicateur', 'Valeur']] + stats_data, colWidths=[6*cm, 4*cm])
    stats_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2D5F3F')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    elements.append(Paragraph('<b>Statistiques</b>', styles['Heading2']))
    elements.append(stats_table)
    elements.append(Spacer(1, 0.8*cm))

    # Taux par membre
    elements.append(Paragraph('<b>Taux et montants par membre</b>', styles['Heading2']))
    elements.append(Spacer(1, 0.3*cm))
    if stats_membres:
        m_headers = [['Membre', 'Payées', 'Total', 'Montant assigné', 'Montant payé', 'Taux (%)']]
        m_data = [[s['nom'], str(s['nb_payees']), str(s['nb_total']),
                   f"{float(s['montant_total']):,.0f}", f"{float(s['montant_paye']):,.0f}",
                   f"{s['taux_cotisation']}%"] for s in stats_membres]
        mt = Table(m_headers + m_data, colWidths=[5*cm, 2*cm, 2*cm, 3*cm, 3*cm, 2*cm])
        mt.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2D5F3F')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        elements.append(mt)
    else:
        elements.append(Paragraph('Aucune cotisation.', styles['Normal']))

    doc.build(elements)
    buf.seek(0)
    return buf
