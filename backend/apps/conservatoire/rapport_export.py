"""
Export des rapports de séances de répétition en PDF, Excel ou CSV.
Inclut toutes les séances, présences, statistiques et infos des kourels
(responsable, maitres de cœur, jewrine).
"""
from .models import SeanceConservatoire, PresenceSeance, Kourel


# ─────────────────────────────── Helpers ────────────────────────────────────

def _get_queryset(date_debut=None, date_fin=None, kourel_id=None):
    qs = SeanceConservatoire.objects.filter(type_seance='repetition')
    if date_debut:
        qs = qs.filter(date_heure__date__gte=date_debut)
    if date_fin:
        qs = qs.filter(date_heure__date__lte=date_fin)
    if kourel_id:
        qs = qs.filter(kourel_id=kourel_id)
    return qs.select_related(
        'kourel',
        'kourel__responsable',
        'kourel__maitre_de_coeur',
        'kourel__maitre_de_coeur_2',
        'kourel__jewrine',
    ).prefetch_related(
        'presences', 'presences__membre', 'khassidas', 'kourel__membres'
    ).order_by('date_heure')


def _get_stats_par_membre(qs):
    """[{membre, nom, kourel, nb_seances_attendues, nb_presents, nb_abs_just, nb_abs_non_just, taux_presence}]"""
    from collections import defaultdict
    seance_ids = list(qs.values_list('id', flat=True))
    membre_seances = defaultdict(set)
    membre_kourel = {}
    for seance in qs.select_related('kourel'):
        kourel = seance.kourel
        if not kourel:
            continue
        for m in kourel.membres.all():
            membre_seances[m.id].add(seance.id)
            if m.id not in membre_kourel:
                membre_kourel[m.id] = (m, kourel.nom)
    presences = PresenceSeance.objects.filter(seance_id__in=seance_ids).select_related('membre')
    from collections import defaultdict
    membre_presents = defaultdict(int)
    membre_abs_just = defaultdict(int)
    membre_abs_non_just = defaultdict(int)
    for p in presences:
        mid = p.membre_id
        if p.statut == 'present':
            membre_presents[mid] += 1
        elif p.statut == 'absent_justifie':
            membre_abs_just[mid] += 1
        else:
            membre_abs_non_just[mid] += 1
    result = []
    for mid, seance_ids_att in membre_seances.items():
        nb_attendues = len(seance_ids_att)
        nb_pres = membre_presents[mid]
        nb_aj = membre_abs_just[mid]
        nb_anj = membre_abs_non_just[mid]
        taux = round(100 * nb_pres / nb_attendues, 1) if nb_attendues else 0
        m, kourel_nom = membre_kourel.get(mid, (None, ''))
        nom = m.get_full_name() if m else f'Membre #{mid}'
        result.append({
            'membre_id': mid, 'membre': m, 'nom': nom, 'kourel': kourel_nom,
            'nb_seances_attendues': nb_attendues, 'nb_presents': nb_pres,
            'nb_abs_just': nb_aj, 'nb_abs_non_just': nb_anj,
            'taux_presence': taux
        })
    return sorted(result, key=lambda x: (-x['taux_presence'], x['nom']))


def _kourel_encadrement(kourel):
    """Retourne un dict avec les noms des encadrants d'un kourel."""
    return {
        'responsable': kourel.responsable.get_full_name() if kourel.responsable else '—',
        'maitre_1': kourel.maitre_de_coeur.get_full_name() if kourel.maitre_de_coeur else '—',
        'maitre_2': kourel.maitre_de_coeur_2.get_full_name() if kourel.maitre_de_coeur_2 else '—',
        'jewrine': kourel.jewrine.get_full_name() if kourel.jewrine else '—',
    }


# ────────────────────────── Excel export ────────────────────────────────────

_GREEN = '2D5F3F'
_GOLD  = 'C9A961'
_LIGHT = 'F4EAD5'
_GREY  = 'DDDDDD'


def export_rapport_excel(date_debut=None, date_fin=None, kourel_id=None):
    """Export Excel multi-feuilles : stats globales, fiche kourels, taux membres, détail séances."""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
        from openpyxl.utils import get_column_letter
    except ImportError:
        return None

    qs = _get_queryset(date_debut, date_fin, kourel_id)
    wb = openpyxl.Workbook()

    # ── Styles communs ──
    def hdr_font(bold=True, white=False, size=10):
        return Font(bold=bold, color='FFFFFF' if white else '000000', size=size)

    def fill(hex_color):
        return PatternFill('solid', fgColor=hex_color)

    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _write_header_row(ws, row_num, headers, bg_color=_GREEN, text_color='FFFFFF', bold=True):
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=row_num, column=col, value=h)
            c.font = Font(bold=bold, color=text_color)
            c.fill = PatternFill('solid', fgColor=bg_color)
            c.border = border
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    def _autofit(ws, col_widths):
        for col, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = w

    periode_str = (
        f"{date_debut.strftime('%d/%m/%Y')} — {date_fin.strftime('%d/%m/%Y')}"
        if date_debut and date_fin else "Toutes les séances créées"
    )

    # ══════════════════════════════════════════════════
    # Feuille 1 : Statistiques globales
    # ══════════════════════════════════════════════════
    ws_stats = wb.active
    ws_stats.title = "Statistiques"

    presences_all = PresenceSeance.objects.filter(seance__in=qs)
    nb_seances = qs.count()
    nb_total = presences_all.count()
    nb_presents = presences_all.filter(statut='present').count()
    nb_abs_just = presences_all.filter(statut='absent_justifie').count()
    nb_abs_non_just = presences_all.filter(statut='absent_non_justifie').count()
    taux_presence = round(100 * nb_presents / nb_total, 1) if nb_total else 0

    ws_stats.merge_cells('A1:C1')
    title_cell = ws_stats.cell(row=1, column=1, value="Rapport des Séances de Répétition — DBM")
    title_cell.font = Font(bold=True, size=14, color='FFFFFF')
    title_cell.fill = fill(_GREEN)
    title_cell.alignment = Alignment(horizontal='center')

    ws_stats.merge_cells('A2:C2')
    period_cell = ws_stats.cell(row=2, column=1, value=f"Période : {periode_str}")
    period_cell.font = Font(italic=True, size=10)
    period_cell.fill = fill(_LIGHT)
    period_cell.alignment = Alignment(horizontal='center')

    _write_header_row(ws_stats, 4, ['Indicateur', 'Valeur', ''], bg_color=_GOLD, text_color='000000')

    stats_rows = [
        ("Nombre total de séances", nb_seances),
        ("Total marquages présence", nb_total),
        ("Présents", nb_presents),
        ("Absents justifiés", nb_abs_just),
        ("Absents non justifiés", nb_abs_non_just),
        (f"Taux de présence global", f"{taux_presence}%"),
    ]
    for i, (label, val) in enumerate(stats_rows, 5):
        ws_stats.cell(row=i, column=1, value=label).border = border
        c = ws_stats.cell(row=i, column=2, value=val)
        c.border = border
        c.alignment = Alignment(horizontal='center')
        if i % 2 == 0:
            for col in [1, 2]:
                ws_stats.cell(row=i, column=col).fill = fill(_GREY)
    _autofit(ws_stats, [32, 18, 5])

    # ══════════════════════════════════════════════════
    # Feuille 2 : Fiche Kourels
    # ══════════════════════════════════════════════════
    ws_kourels = wb.create_sheet("Kourels")
    _write_header_row(ws_kourels, 1, [
        'Kourel', 'Responsable', '1er Maître de cœur', '2ème Maître de cœur',
        'Jewrine', 'Nb membres', 'Nb séances'
    ])
    kourel_qs = Kourel.objects.select_related(
        'responsable', 'maitre_de_coeur', 'maitre_de_coeur_2', 'jewrine'
    ).prefetch_related('membres')
    if kourel_id:
        kourel_qs = kourel_qs.filter(pk=kourel_id)

    seances_par_kourel = {}
    for s in qs:
        seances_par_kourel[s.kourel_id] = seances_par_kourel.get(s.kourel_id, 0) + 1

    for row_i, k in enumerate(kourel_qs.order_by('ordre', 'nom'), 2):
        enc = _kourel_encadrement(k)
        row_data = [
            k.nom, enc['responsable'], enc['maitre_1'], enc['maitre_2'],
            enc['jewrine'], k.membres.count(), seances_par_kourel.get(k.pk, 0)
        ]
        bg = _LIGHT if row_i % 2 == 0 else 'FFFFFF'
        for col, val in enumerate(row_data, 1):
            c = ws_kourels.cell(row=row_i, column=col, value=val)
            c.border = border
            c.fill = fill(bg)
            c.alignment = Alignment(horizontal='center' if col > 1 else 'left')
    _autofit(ws_kourels, [22, 22, 22, 22, 22, 12, 12])

    # ══════════════════════════════════════════════════
    # Feuille 3 : Taux de présence par membre
    # ══════════════════════════════════════════════════
    stats_membres = _get_stats_par_membre(qs)
    ws_membres = wb.create_sheet("Taux par membre")
    _write_header_row(ws_membres, 1, [
        'Membre', 'Kourel', 'Séances attendues',
        'Présents', 'Abs. justifiés', 'Abs. non justifiés', 'Taux présence (%)'
    ])
    for i, s in enumerate(stats_membres, 2):
        taux_val = s['taux_presence']
        row_data = [
            s['nom'], s['kourel'], s['nb_seances_attendues'],
            s['nb_presents'], s['nb_abs_just'], s['nb_abs_non_just'], f"{taux_val}%"
        ]
        # Couleur selon taux
        if taux_val >= 80:
            row_bg = 'E8F5E9'
        elif taux_val >= 50:
            row_bg = 'FFF9E6'
        else:
            row_bg = 'FFEBEE'
        for col, val in enumerate(row_data, 1):
            c = ws_membres.cell(row=i, column=col, value=val)
            c.border = border
            c.fill = fill(row_bg)
            c.alignment = Alignment(horizontal='center' if col > 1 else 'left')
    _autofit(ws_membres, [28, 20, 16, 12, 16, 18, 16])

    # ══════════════════════════════════════════════════
    # Feuille 4 : Détail des séances
    # ══════════════════════════════════════════════════
    ws = wb.create_sheet("Détail séances")
    taux_par_membre = {s['membre_id']: s['taux_presence'] for s in stats_membres}
    _write_header_row(ws, 1, [
        'Date', 'Heure début', 'Heure fin', 'Lieu', 'Kourel',
        'Responsable', '1er MC', '2ème MC', 'Jewrine',
        'Khassidas (nom, dathie, portion)', 'Membre', 'Statut', 'Taux présence (%)', 'Remarque'
    ])

    row = 2
    for seance in qs:
        kourel = seance.kourel
        enc = _kourel_encadrement(kourel) if kourel else {
            'responsable': '—', 'maitre_1': '—', 'maitre_2': '—', 'jewrine': '—'
        }
        khass_str = ' | '.join(
            f"{k.nom_khassida} ({k.dathie})" + (f" - {k.khassida_portion}" if k.khassida_portion else "")
            for k in seance.khassidas.all()
        ) or '—'
        date_str = seance.date_heure.strftime('%d/%m/%Y') if seance.date_heure else ''
        heure_debut = seance.date_heure.strftime('%H:%M') if seance.date_heure else ''
        heure_fin_str = seance.heure_fin.strftime('%H:%M') if seance.heure_fin else ''
        kourel_nom = kourel.nom if kourel else ''

        presences = list(seance.presences.all())
        base_row = [
            date_str, heure_debut, heure_fin_str, seance.lieu or '',
            kourel_nom, enc['responsable'], enc['maitre_1'], enc['maitre_2'],
            enc['jewrine'], khass_str
        ]
        if not presences:
            row_data = base_row + ['—', '—', '—', '']
            for col, val in enumerate(row_data, 1):
                c = ws.cell(row=row, column=col, value=val)
                c.border = border
                c.fill = fill(_LIGHT)
            row += 1
        else:
            for idx_p, p in enumerate(presences):
                taux = f"{taux_par_membre.get(p.membre_id, 0)}%" if p.membre_id else '—'
                statut_display = p.get_statut_display()
                row_data = base_row + [
                    p.membre.get_full_name() if p.membre else '',
                    statut_display, taux, p.remarque or ''
                ]
                bg = _LIGHT if idx_p % 2 == 0 else 'FFFFFF'
                for col, val in enumerate(row_data, 1):
                    c = ws.cell(row=row, column=col, value=val)
                    c.border = border
                    c.fill = fill(bg)
                row += 1

    _autofit(ws, [12, 10, 10, 18, 18, 18, 18, 18, 18, 30, 24, 18, 14, 24])

    from io import BytesIO
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ────────────────────────── PDF export ──────────────────────────────────────

def export_rapport_pdf(date_debut=None, date_fin=None, kourel_id=None):
    """Export PDF complet : stats, fiche kourels, taux membres, détail séances."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, HRFlowable
        )
    except ImportError:
        return None

    qs = _get_queryset(date_debut, date_fin, kourel_id)
    from io import BytesIO
    buf = BytesIO()

    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        rightMargin=1.5*cm, leftMargin=1.5*cm,
        topMargin=2*cm, bottomMargin=2*cm
    )
    styles = getSampleStyleSheet()
    GREEN = colors.HexColor('#2D5F3F')
    GOLD  = colors.HexColor('#C9A961')
    LIGHT = colors.HexColor('#F4EAD5')
    RED   = colors.HexColor('#CC3333')
    ORANGE = colors.HexColor('#E6A817')

    h1_style = ParagraphStyle('H1', parent=styles['Heading1'], textColor=GREEN, fontSize=16, spaceAfter=4)
    h2_style = ParagraphStyle('H2', parent=styles['Heading2'], textColor=GREEN, fontSize=12, spaceAfter=4)
    small = ParagraphStyle('Small', parent=styles['Normal'], fontSize=8)

    def tbl_style(header_color=GREEN, stripe=LIGHT):
        return TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), header_color),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 7),
            ('TOPPADDING', (0, 0), (-1, 0), 7),
            ('GRID', (0, 0), (-1, -1), 0.4, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, stripe]),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('TOPPADDING', (0, 1), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ])

    elements = []
    periode_str = (
        f"{date_debut.strftime('%d/%m/%Y')} — {date_fin.strftime('%d/%m/%Y')}"
        if date_debut and date_fin else "Toutes les séances créées"
    )

    try:
        from utils.pdf_header import build_pdf_header
        elements.extend(build_pdf_header("Rapport des séances de répétition", periode_str))
    except Exception:
        elements.append(Paragraph("Rapport des séances de répétition", h1_style))
        elements.append(Paragraph(f"Période : {periode_str}", styles['Normal']))
        elements.append(Spacer(1, 0.5*cm))

    # ── Statistiques globales ──
    presences_all = list(PresenceSeance.objects.filter(seance__in=qs))
    nb_total = len(presences_all)
    nb_presents = sum(1 for p in presences_all if p.statut == 'present')
    nb_abs_just = sum(1 for p in presences_all if p.statut == 'absent_justifie')
    nb_abs_non_just = sum(1 for p in presences_all if p.statut == 'absent_non_justifie')
    taux_presence = round(100 * nb_presents / nb_total, 1) if nb_total else 0

    elements.append(Paragraph('Statistiques globales', h2_style))
    stats_data = [['Indicateur', 'Valeur']] + [
        ['Nombre de séances', str(qs.count())],
        ['Total marquages présence', str(nb_total)],
        ['Présents', str(nb_presents)],
        ['Absents justifiés', str(nb_abs_just)],
        ['Absents non justifiés', str(nb_abs_non_just)],
        ['Taux de présence global', f'{taux_presence}%'],
    ]
    stats_table = Table(stats_data, colWidths=[9*cm, 5*cm])
    stats_table.setStyle(tbl_style(header_color=GOLD))
    elements.append(stats_table)
    elements.append(Spacer(1, 0.8*cm))

    # ── Fiche Kourels ──
    elements.append(Paragraph('Encadrement des Kourels', h2_style))
    elements.append(Spacer(1, 0.3*cm))
    kourel_qs = Kourel.objects.select_related(
        'responsable', 'maitre_de_coeur', 'maitre_de_coeur_2', 'jewrine'
    ).prefetch_related('membres').order_by('ordre', 'nom')
    if kourel_id:
        kourel_qs = kourel_qs.filter(pk=kourel_id)

    kourel_headers = [['Kourel', 'Responsable', '1er MC', '2ème MC', 'Jewrine', 'Membres']]
    kourel_rows = []
    for k in kourel_qs:
        enc = _kourel_encadrement(k)
        kourel_rows.append([k.nom, enc['responsable'], enc['maitre_1'], enc['maitre_2'], enc['jewrine'], str(k.membres.count())])
    if kourel_rows:
        kt = Table(kourel_headers + kourel_rows, colWidths=[3.5*cm, 3.5*cm, 3*cm, 3*cm, 3*cm, 2*cm])
        kt.setStyle(tbl_style())
        elements.append(kt)
    else:
        elements.append(Paragraph('Aucun kourel trouvé.', styles['Normal']))
    elements.append(Spacer(1, 0.8*cm))

    # ── Taux par membre ──
    stats_membres = _get_stats_par_membre(qs)
    elements.append(Paragraph('Taux de présence par membre', h2_style))
    elements.append(Spacer(1, 0.3*cm))
    if stats_membres:
        m_headers = [['Membre', 'Kourel', 'Attendues', 'Présents', 'Abs. just.', 'Abs. non just.', 'Taux (%)']]
        m_data = [
            [s['nom'], s['kourel'], str(s['nb_seances_attendues']), str(s['nb_presents']),
             str(s['nb_abs_just']), str(s['nb_abs_non_just']), f"{s['taux_presence']}%"]
            for s in stats_membres
        ]
        mt = Table(m_headers + m_data, colWidths=[4.5*cm, 3.5*cm, 2*cm, 2*cm, 2*cm, 2.5*cm, 2*cm])
        style = tbl_style()
        # Coloration conditionnelle du taux
        for i, s in enumerate(stats_membres, 1):
            taux = s['taux_presence']
            color = colors.HexColor('#E8F5E9') if taux >= 80 else (colors.HexColor('#FFF9E6') if taux >= 50 else colors.HexColor('#FFEBEE'))
            style.add('BACKGROUND', (6, i), (6, i), color)
        mt.setStyle(style)
        elements.append(mt)
    else:
        elements.append(Paragraph('Aucun membre concerné.', styles['Normal']))
    elements.append(Spacer(1, 0.8*cm))

    # ── Détail des séances ──
    elements.append(PageBreak())
    elements.append(Paragraph('Détail des séances', h2_style))
    elements.append(Spacer(1, 0.3*cm))

    taux_par_membre = {s['membre_id']: s['taux_presence'] for s in stats_membres}
    for seance in qs:
        kourel = seance.kourel
        enc = _kourel_encadrement(kourel) if kourel else {
            'responsable': '—', 'maitre_1': '—', 'maitre_2': '—', 'jewrine': '—'
        }
        khass_str = ', '.join(
            f"{k.nom_khassida} ({k.dathie})" + (f" - {k.khassida_portion}" if k.khassida_portion else "")
            for k in seance.khassidas.all()
        ) or '—'
        date_str = seance.date_heure.strftime('%d/%m/%Y %H:%M') if seance.date_heure else ''
        heure_fin_str = seance.heure_fin.strftime('%H:%M') if seance.heure_fin else '—'

        # Entête séance
        seance_info = Table([
            ['Date/Heure', 'Heure fin', 'Lieu', 'Kourel'],
            [date_str, heure_fin_str, seance.lieu or '—', kourel.nom if kourel else '—'],
            ['Responsable', '1er MC', '2ème MC', 'Jewrine'],
            [enc['responsable'], enc['maitre_1'], enc['maitre_2'], enc['jewrine']],
        ], colWidths=[4*cm, 3*cm, 4.5*cm, 4.5*cm])
        seance_info.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), GREEN),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 1), (-1, 1), LIGHT),
            ('BACKGROUND', (0, 2), (-1, 2), GOLD),
            ('FONTNAME', (0, 2), (-1, 2), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 3), (-1, 3), colors.HexColor('#FAF3E0')),
            ('GRID', (0, 0), (-1, -1), 0.4, colors.grey),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(seance_info)

        if khass_str != '—':
            elements.append(Spacer(1, 0.15*cm))
            elements.append(Paragraph(f"<b>Khassidas :</b> {khass_str}", small))

        presences = list(seance.presences.all())
        if presences:
            elements.append(Spacer(1, 0.2*cm))
            ph = [['Membre', 'Statut', 'Taux présence', 'Remarque']]
            for p in presences:
                taux = f"{taux_par_membre.get(p.membre_id, 0)}%" if p.membre_id else '—'
                ph.append([
                    p.membre.get_full_name() if p.membre else '',
                    p.get_statut_display(), taux, p.remarque or ''
                ])
            pt = Table(ph, colWidths=[5*cm, 3.5*cm, 2.5*cm, 5*cm])
            pt_style = tbl_style(header_color=GOLD)
            # Coloriser statut
            for i, p in enumerate(presences, 1):
                if p.statut == 'present':
                    pt_style.add('TEXTCOLOR', (1, i), (1, i), GREEN)
                elif p.statut == 'absent_non_justifie':
                    pt_style.add('TEXTCOLOR', (1, i), (1, i), RED)
                else:
                    pt_style.add('TEXTCOLOR', (1, i), (1, i), ORANGE)
            pt.setStyle(pt_style)
            elements.append(pt)
        elements.append(Spacer(1, 0.6*cm))

    doc.build(elements)
    buf.seek(0)
    return buf


# ────────────────────────── CSV export ──────────────────────────────────────

def export_rapport_csv(date_debut=None, date_fin=None, kourel_id=None):
    """Export CSV complet des séances de répétition, avec infos encadrement kourel."""
    import csv
    from io import StringIO

    qs = _get_queryset(date_debut, date_fin, kourel_id)
    buf = StringIO()
    writer = csv.writer(buf, delimiter=';')

    periode_str = (
        f"{date_debut.strftime('%d/%m/%Y')} — {date_fin.strftime('%d/%m/%Y')}"
        if date_debut and date_fin else "Toutes les séances créées"
    )
    writer.writerow([f'Rapport séances de répétition DBM — Période : {periode_str}'])

    presences_all = list(PresenceSeance.objects.filter(seance__in=qs))
    nb_total = len(presences_all)
    nb_presents = sum(1 for p in presences_all if p.statut == 'present')
    taux_presence = round(100 * nb_presents / nb_total, 1) if nb_total else 0
    writer.writerow([f'Séances: {qs.count()} | Présences: {nb_total} | Présents: {nb_presents} | Taux: {taux_presence}%'])
    writer.writerow([])

    # Section kourels
    writer.writerow(['=== ENCADREMENT DES KOURELS ==='])
    writer.writerow(['Kourel', 'Responsable', '1er Maître de cœur', '2ème Maître de cœur', 'Jewrine', 'Nb membres'])
    kourel_qs = Kourel.objects.select_related(
        'responsable', 'maitre_de_coeur', 'maitre_de_coeur_2', 'jewrine'
    ).prefetch_related('membres').order_by('ordre', 'nom')
    if kourel_id:
        kourel_qs = kourel_qs.filter(pk=kourel_id)
    for k in kourel_qs:
        enc = _kourel_encadrement(k)
        writer.writerow([k.nom, enc['responsable'], enc['maitre_1'], enc['maitre_2'], enc['jewrine'], k.membres.count()])
    writer.writerow([])

    # Section taux par membre
    stats_membres = _get_stats_par_membre(qs)
    writer.writerow(['=== TAUX DE PRESENCE PAR MEMBRE ==='])
    writer.writerow(['Membre', 'Kourel', 'Séances attendues', 'Présents', 'Abs. justifiés', 'Abs. non justifiés', 'Taux (%)'])
    for s in stats_membres:
        writer.writerow([s['nom'], s['kourel'], s['nb_seances_attendues'], s['nb_presents'],
                        s['nb_abs_just'], s['nb_abs_non_just'], f"{s['taux_presence']}%"])
    writer.writerow([])

    # Section détail séances
    writer.writerow(['=== DETAIL DES SEANCES ==='])
    writer.writerow([
        'Date', 'Heure début', 'Heure fin', 'Lieu', 'Kourel',
        'Responsable', '1er MC', '2ème MC', 'Jewrine',
        'Khassidas (nom, dathie, portion)', 'Membre', 'Statut', 'Taux présence (%)', 'Remarque'
    ])
    taux_par_membre = {s['membre_id']: s['taux_presence'] for s in stats_membres}
    for seance in qs:
        kourel = seance.kourel
        enc = _kourel_encadrement(kourel) if kourel else {
            'responsable': '—', 'maitre_1': '—', 'maitre_2': '—', 'jewrine': '—'
        }
        khass_str = ' | '.join(
            f"{k.nom_khassida} ({k.dathie})" + (f" - {k.khassida_portion}" if k.khassida_portion else "")
            for k in seance.khassidas.all()
        ) or '—'
        date_str = seance.date_heure.strftime('%d/%m/%Y') if seance.date_heure else ''
        heure_debut = seance.date_heure.strftime('%H:%M') if seance.date_heure else ''
        heure_fin_str = seance.heure_fin.strftime('%H:%M') if seance.heure_fin else ''
        kourel_nom = kourel.nom if kourel else ''
        base = [date_str, heure_debut, heure_fin_str, seance.lieu or '',
                kourel_nom, enc['responsable'], enc['maitre_1'], enc['maitre_2'], enc['jewrine'], khass_str]

        presences = list(seance.presences.all())
        if not presences:
            writer.writerow(base + ['—', '—', '—', ''])
        else:
            for p in presences:
                taux = f"{taux_par_membre.get(p.membre_id, 0)}%" if p.membre_id else '—'
                writer.writerow(base + [
                    p.membre.get_full_name() if p.membre else '',
                    p.get_statut_display(), taux, p.remarque or ''
                ])

    from io import BytesIO
    out = BytesIO(buf.getvalue().encode('utf-8-sig'))
    out.seek(0)
    return out


# ────────────────────────── Export fiche membres kourel ─────────────────────

def export_membres_kourel_excel(kourel_id):
    """Export Excel de la liste des membres d'un kourel avec leurs infos."""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        return None

    try:
        kourel = Kourel.objects.select_related(
            'responsable', 'maitre_de_coeur', 'maitre_de_coeur_2', 'jewrine'
        ).prefetch_related('membres').get(pk=kourel_id)
    except Kourel.DoesNotExist:
        return None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Membres — {kourel.nom[:20]}"

    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def fill(hex_color):
        return PatternFill('solid', fgColor=hex_color)

    # Entête kourel
    ws.merge_cells('A1:G1')
    ws['A1'] = f"FICHE KOUREL : {kourel.nom.upper()}"
    ws['A1'].font = Font(bold=True, size=14, color='FFFFFF')
    ws['A1'].fill = fill(_GREEN)
    ws['A1'].alignment = Alignment(horizontal='center')

    enc = _kourel_encadrement(kourel)
    info_rows = [
        ('Responsable', enc['responsable']),
        ('1er Maître de cœur', enc['maitre_1']),
        ('2ème Maître de cœur', enc['maitre_2']),
        ('Jewrine', enc['jewrine']),
        ('Nombre de membres', str(kourel.membres.count())),
    ]
    for i, (label, val) in enumerate(info_rows, 2):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws.cell(row=i, column=1).fill = fill(_LIGHT)
        ws.cell(row=i, column=2, value=val)
        ws.cell(row=i, column=1).border = border
        ws.cell(row=i, column=2).border = border

    # En-tête membres
    header_row = len(info_rows) + 3
    headers = ['#', 'Nom complet', 'Téléphone', 'Email', 'Profession', 'Catégorie', 'Cellule']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=header_row, column=col, value=h)
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = fill(_GOLD)
        c.border = border
        c.alignment = Alignment(horizontal='center')

    for i, membre in enumerate(kourel.membres.all().order_by('last_name', 'first_name'), 1):
        row_num = header_row + i
        row_data = [
            i,
            membre.get_full_name(),
            membre.telephone or '—',
            membre.email or '—',
            membre.profession or '—',
            membre.get_categorie_display() if hasattr(membre, 'get_categorie_display') else membre.categorie or '—',
            membre.get_cellule_display() if hasattr(membre, 'get_cellule_display') else membre.cellule or '—',
        ]
        bg = _LIGHT if i % 2 == 0 else 'FFFFFF'
        for col, val in enumerate(row_data, 1):
            c = ws.cell(row=row_num, column=col, value=val)
            c.border = border
            c.fill = fill(bg)
            c.alignment = Alignment(horizontal='center' if col == 1 else 'left')

    col_widths = [5, 28, 16, 28, 20, 16, 16]
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    from io import BytesIO
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_membres_kourel_pdf(kourel_id):
    """Export PDF de la fiche membres d'un kourel."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    except ImportError:
        return None

    try:
        kourel = Kourel.objects.select_related(
            'responsable', 'maitre_de_coeur', 'maitre_de_coeur_2', 'jewrine'
        ).prefetch_related('membres').get(pk=kourel_id)
    except Kourel.DoesNotExist:
        return None

    from io import BytesIO
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    GREEN = colors.HexColor('#2D5F3F')
    GOLD  = colors.HexColor('#C9A961')
    LIGHT = colors.HexColor('#F4EAD5')

    h1 = ParagraphStyle('H1', parent=styles['Heading1'], textColor=GREEN, fontSize=15)
    h2 = ParagraphStyle('H2', parent=styles['Heading2'], textColor=GREEN, fontSize=11)

    elements = []
    elements.append(Paragraph(f"Fiche Kourel : {kourel.nom}", h1))
    elements.append(Spacer(1, 0.4*cm))

    enc = _kourel_encadrement(kourel)
    enc_data = [
        ['Responsable', enc['responsable']],
        ['1er Maître de cœur', enc['maitre_1']],
        ['2ème Maître de cœur', enc['maitre_2']],
        ['Jewrine', enc['jewrine']],
        ['Nombre de membres', str(kourel.membres.count())],
    ]
    enc_tbl = Table(enc_data, colWidths=[6*cm, 11*cm])
    enc_tbl.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), LIGHT),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.grey),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    elements.append(enc_tbl)
    elements.append(Spacer(1, 0.6*cm))

    elements.append(Paragraph("Liste des membres", h2))
    elements.append(Spacer(1, 0.3*cm))

    membres = list(kourel.membres.all().order_by('last_name', 'first_name'))
    if membres:
        m_headers = [['#', 'Nom complet', 'Téléphone', 'Email', 'Catégorie']]
        m_data = [
            [str(i), m.get_full_name(), m.telephone or '—', m.email or '—',
             m.get_categorie_display() if hasattr(m, 'get_categorie_display') else m.categorie or '—']
            for i, m in enumerate(membres, 1)
        ]
        mt = Table(m_headers + m_data, colWidths=[1*cm, 5*cm, 3.5*cm, 5*cm, 3*cm])
        mt.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), GOLD),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.4, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, LIGHT]),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(mt)
    else:
        elements.append(Paragraph('Aucun membre enregistré dans ce kourel.', styles['Normal']))

    doc.build(elements)
    buf.seek(0)
    return buf
