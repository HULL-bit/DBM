"""Export du journal de sécurité/traçabilité (JournalAudit) en Excel ou PDF."""
from io import BytesIO


def _lignes(qs):
    return [
        [
            e.date.strftime('%d/%m/%Y %H:%M'),
            e.utilisateur.get_full_name() if e.utilisateur else '—',
            e.get_action_display(),
            e.rubrique or '—',
            e.description or e.objet_repr or '—',
            e.adresse_ip or '—',
            'OK' if e.succes else 'Échec',
        ]
        for e in qs
    ]


def export_audit_excel(qs):
    try:
        import openpyxl
        from openpyxl.styles import Font, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Journal de sécurité"
    header_font = Font(bold=True)
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.cell(row=1, column=1, value="Journal de sécurité — Daara Barakatul Mahaahidi").font = Font(bold=True, size=14)

    headers = ['Date', 'Utilisateur', 'Action', 'Rubrique', 'Description', 'Adresse IP', 'Résultat']
    row0 = 3
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row0, column=col, value=h)
        c.font = header_font
        c.border = border

    for i, ligne in enumerate(_lignes(qs), row0 + 1):
        for col, val in enumerate(ligne, 1):
            ws.cell(row=i, column=col, value=val).border = border

    largeurs = [18, 24, 22, 16, 45, 16, 10]
    for col, w in enumerate(largeurs, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_audit_pdf(qs):
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
    except ImportError:
        return None

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), rightMargin=1.5 * cm, leftMargin=1.5 * cm, topMargin=1.5 * cm, bottomMargin=1.5 * cm)
    elements = []

    from utils.pdf_header import build_pdf_header
    elements.extend(build_pdf_header("Journal de sécurité", "Traçabilité des actions sensibles"))

    data = [['Date', 'Utilisateur', 'Action', 'Rubrique', 'Description', 'IP', 'Résultat']]
    data.extend(_lignes(qs))

    table = Table(data, colWidths=[3 * cm, 3.5 * cm, 3.5 * cm, 2.5 * cm, 8 * cm, 2.5 * cm, 1.8 * cm], repeatRows=1)
    style = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2D5F3F')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]
    table.setStyle(TableStyle(style))
    elements.append(table)

    doc.build(elements)
    buf.seek(0)
    return buf
